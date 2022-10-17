import argparse

from openpyxl import load_workbook


class Node(object):
    """" Generic tree node."""

    def __init__(self, name='root', children=None):
        self.name = name
        self.children = []
        if children is not None:
            for child in children:
                if child:
                    self.add_child(child)

    def __repr__(self):
        return "{} {}".format(self.name, self.children)

    @staticmethod
    def to_tree(node_list):
        root = None
        for i in range(len(node_list))[::-1]:
            cur = Node(node_list[i], [root])
            root = cur

        return root

    def add_child(self, node):
        assert isinstance(node, Node)
        self.children.append(node)

    def is_child(self, node):
        for child in self.children:
            if node.name == child.name:
                return True
        return False

    def insert(self, node):
        """
        insert
        """
        assert isinstance(node, Node)
        # if it exists, go one-level down and compare again
        if self.is_child(node):
            self.insert(node.children[0])
        else:
            self.children.append(node)


class ExcelTreeBuilder:
    def build_from_worksheet(self, worksheet):

        tree = Node("AutoEver")

        id_row = 0
        for row in worksheet.iter_rows():
            if row:

                id_col = 0
                values = list()

                for cell in row:
                    if cell and cell.value:
                        val = cell.value
                    else:
                        val = "-"

                    values.append(val)

                    id_col += 1

                values_tree = Node.to_tree(values)
                print(values_tree)

            id_row += 1

    def build(self, path):
        # tree = Node.to_tree(["a", "b", "c"])
        # tree = Node('a', [Node('b', [Node('c'), Node('d')])])
        # print(tree)

        workbook = load_workbook(path)

        for ws in workbook.worksheets:
            print(ws.title)
            if ws.title == "Sheet1":
                ws = workbook[ws.title]
                self.build_from_worksheet(ws)

            # workbook.save(path[:-5] + "-tr" + path[-5:])


if __name__ == "__main__":
    # Create the parser
    arg_parser = argparse.ArgumentParser()

    # Add arguments
    arg_parser.add_argument('--file_path', help='file path')

    # Execute the parse_args() method
    args = arg_parser.parse_args()
    file_path = args.file_path

    print("file is " + file_path)

    builder = ExcelTreeBuilder()
    builder.build(file_path)
    exit(0)
