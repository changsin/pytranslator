from openpyxl import load_workbook


class ExcelHandler:
    def __init__(self, translator):
        self.translator = translator

    def translate_payload(self, worksheet):

        id_row = 0
        for row in worksheet.iter_rows():
            if row:

                id_col = 0
                row_str = ""
                for cell in row:
                    if cell and cell.value and str(cell.value).strip():
                        to_translate = str(cell.value).strip()
                        print(to_translate)
                        translated_text = self.translator.translate(to_translate)
                        row_str = row_str + "\t: {}->{}".format(cell.value, translated_text)
                        cell.value = translated_text

                    id_col += 1

                if len(row_str) > 0:
                    print(str(id_row) + row_str + "\t")

            id_row += 1

    def translate(self, path):
        workbook = load_workbook(path)

        for ws in workbook.worksheets:
            ws = workbook[ws.title]
            self.translate_payload(ws)

            workbook.save(path[:-5] + "-tr" + path[-5:])
