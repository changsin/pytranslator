import json
import argparse

from json_utils import from_file, to_file
from openpyxl import Workbook
from openpyxl import load_workbook
from translators.translator_azure import TranslatorAzure
from translators.translator_google import TranslatorGoogle
from translators.translator_papago import TranslatorPapago


class ExcelTranslator:
    def __init__(self, path, translator):
        self.path = path
        self.workbook = load_workbook(path)
        self.translator = translator
        self.out_workbook = Workbook()

    def get_worksheet(self, name):
        return self.workbook[name]

    def translate_data(self, worksheet):

        id_row = 0
        for row in worksheet.iter_rows():
            if row:

                id_col = 0
                row_str = ""
                for cell in row:
                    if cell and cell.value and str(cell.value).strip():
                        translated_text = self.translator.translate(str(cell.value).strip())
                        row_str = row_str + "\t: {}->{}".format(cell.value, translated_text)
                        cell.value = translated_text

                    id_col += 1

                if len(row_str) > 0:
                    print(str(id_row) + row_str + "\t")

            id_row += 1

        self.workbook.save(self.path[:-5] + "-tr" + self.path[-5:])


if __name__ == "__main__":
    # Create the parser
    arg_parser = argparse.ArgumentParser()

    # Add arguments
    arg_parser.add_argument('--path', help='file path')
    arg_parser.add_argument('--target', help='target language')
    arg_parser.add_argument('--source', help='source language')
    arg_parser.add_argument('--dictionary_in', help='input dictionary path')
    arg_parser.add_argument('--dictionary_out', help='output dictionary path')

    # Execute the parse_args() method
    args = arg_parser.parse_args()

    # List of supported languages
    # print(googletrans.LANGUAGES)

    file_path = args.path
    to_language = args.target
    from_language = args.source
    dictionary_in = from_file(args.dictionary_in)
    dictionary_out = from_file(args.dictionary_out)

    # translator = TranslatorAzure(to_language="en", from_language="ko", dictionary=dictionary)
    translator = TranslatorGoogle(to_language=to_language, from_language=from_language, dictionary=dictionary_in)
    # translator = TranslatorPapago(to_language="en", from_language="ko", dictionary=dictionary)

    excel_translator = ExcelTranslator(file_path, translator)

    for ws in excel_translator.workbook.worksheets:
        ws = excel_translator.get_worksheet(ws.title)
        excel_translator.translate_data(ws)

    json_dump = json.dumps(translator.dictionary)
    to_file(dictionary_out, json_dump)
